import os
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from app.config import settings
from app.models.travel_models import TravelInquiryData, TravelQuoteData, InquiryComplexity
from app.utils.logger import get_logger
from app.utils.exceptions import ExcelServiceError

logger = get_logger(__name__)

class EnhancedExcelQuoteGenerator:
    """Enhanced service for generating Excel travel quotes with high accuracy"""
    
    def __init__(self):
        self.template_path = os.path.join(settings.template_path, "travel_quote_template.xlsx")
        self.output_path = settings.file_storage_path
        self._ensure_directories()
    
    def _ensure_directories(self):
        """Ensure required directories exist"""
        os.makedirs(self.output_path, exist_ok=True)
        os.makedirs(settings.template_path, exist_ok=True)
    
    async def generate_quote(self, inquiry: TravelInquiryData, quote_data: TravelQuoteData) -> str:
        """Generate Excel quote based on inquiry complexity"""
        try:
            # Create new workbook
            wb = Workbook()
            
            # Generate sheets based on inquiry complexity
            if inquiry.inquiry_complexity == InquiryComplexity.COMPLEX:
                self._create_complex_summary_sheet(wb, inquiry, quote_data)
                self._create_destination_breakdown_sheet(wb, inquiry, quote_data)
                self._create_complex_pricing_sheet(wb, quote_data)
                self._create_detailed_itinerary_sheet(wb, quote_data)
            else:
                self._create_simple_summary_sheet(wb, inquiry, quote_data)
                self._create_simple_pricing_sheet(wb, quote_data)
                self._create_simple_itinerary_sheet(wb, quote_data)
            
            # Always create terms sheet
            self._create_enhanced_terms_sheet(wb, quote_data)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            complexity = inquiry.inquiry_complexity.value.title()
            filename = f"travel_quote_{complexity}_{quote_data.quote_id}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_path, filename)
            
            # Save the workbook
            wb.save(filepath)
            
            logger.info(f"Enhanced Excel quote generated: {filename}")
            return filepath
            
        except Exception as e:
            logger.error(f"Failed to generate Excel quote: {e}")
            raise ExcelServiceError(f"Excel generation failed: {e}")
    
    def _create_complex_summary_sheet(self, wb: Workbook, inquiry: TravelInquiryData, quote_data: TravelQuoteData):
        """Create summary sheet for complex inquiries"""
        ws = wb.active
        ws.title = "Travel Summary"
        
        # Styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_font = Font(bold=True, size=12, color="000080")
        
        # Main header
        ws.merge_cells('A1:H1')
        ws['A1'] = "COMPREHENSIVE TRAVEL QUOTATION"
        ws['A1'].font = Font(bold=True, size=16, color="000080")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        
        # Quote information section
        self._add_section_header(ws, row, "QUOTE INFORMATION", header_font, header_fill)
        row += 2
        
        quote_info = [
            ("Quote ID:", quote_data.quote_id),
            ("Quote Type:", "Complex Multi-Destination"),
            ("Generated Date:", datetime.now().strftime("%Y-%m-%d")),
            ("Valid Until:", quote_data.valid_until.strftime("%Y-%m-%d") if quote_data.valid_until else "30 days"),
            ("Number of Options:", quote_data.summary.get('number_of_options', len(quote_data.pricing_options)))
        ]
        
        for label, value in quote_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Traveler details section
        self._add_section_header(ws, row, "TRAVELER DETAILS", header_font, header_fill)
        row += 2
        
        traveler_details = [
            ("Total Travelers:", inquiry.traveler_info.total or "Not specified"),
            ("Adults:", inquiry.traveler_info.adults or "Not specified"),
            ("Children:", inquiry.traveler_info.children or "Not specified"),
            ("Couples:", inquiry.traveler_info.couples or "Not specified"),
            ("Singles:", inquiry.traveler_info.singles or "Not specified"),
            ("Visa Required For:", f"{inquiry.traveler_info.visa_required_count} travelers" if inquiry.traveler_info.visa_required_count else "Not specified"),
        ]
        
        for label, value in traveler_details:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Travel overview section
        self._add_section_header(ws, row, "TRAVEL OVERVIEW", header_font, header_fill)
        row += 2
        
        travel_overview = [
            ("Destinations:", ", ".join(inquiry.destinations or []) if inquiry.destinations else "Not specified"),
            ("Travel Dates:", self._format_travel_dates(inquiry.travel_dates)),
            ("Total Duration:", f"{(inquiry.duration or {}).get('total_days', 'Not specified')} days, {(inquiry.duration or {}).get('total_nights', 'Not specified')} nights" if inquiry.duration else "Not specified"),
            ("Departure City:", inquiry.departure_city or "Not specified"),
            ("Budget per Person:", f"₹{inquiry.budget_per_person:,.2f}" if inquiry.budget_per_person else "Not specified"),
        ]
        
        for label, value in travel_overview:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Services required section
        self._add_section_header(ws, row, "SERVICES REQUIRED", header_font, header_fill)
        row += 2
        
        services = [
            ("Visa Assistance:", "Yes" if inquiry.visa_assistance else "No"),
            ("Travel Insurance:", "Yes" if inquiry.insurance_required else "No"),
            ("Flight Booking:", "Yes" if inquiry.flight_required else "No"),
            ("Airport Transfers:", "Yes" if inquiry.airport_transfers else "No"),
            ("Guide Services:", ", ".join(inquiry.guide_language_preferences or []) if inquiry.guide_language_preferences else "As required"),
        ]
        
        for label, value in services:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Special requirements section
        if inquiry.accessibility_requirements or inquiry.dietary_restrictions:
            self._add_section_header(ws, row, "SPECIAL REQUIREMENTS", header_font, header_fill)
            row += 2
            
            if inquiry.accessibility_requirements:
                ws[f'A{row}'] = "Accessibility:"
                ws[f'B{row}'] = ", ".join(inquiry.accessibility_requirements or [])
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            if inquiry.dietary_restrictions:
                ws[f'A{row}'] = "Dietary:"
                ws[f'B{row}'] = ", ".join(inquiry.dietary_restrictions or [])
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        # Apply styling
        self._apply_sheet_styling(ws)
    
    def _create_destination_breakdown_sheet(self, wb: Workbook, inquiry: TravelInquiryData, quote_data: TravelQuoteData):
        """Create destination-wise breakdown sheet for complex inquiries"""
        ws = wb.create_sheet("Destination Details")
        
        # Header
        ws.merge_cells('A1:G1')
        ws['A1'] = "DESTINATION-WISE BREAKDOWN"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        
        for i, dest_detail in enumerate(inquiry.destination_details or [], 1):
            # Destination header
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = f"DESTINATION {i}: {dest_detail.destination_name.upper()}"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            row += 2
            
            # Destination details
            dest_info = [
                ("Duration:", f"{dest_detail.nights} nights" if dest_detail.nights else "Not specified"),
                ("Accommodation:", self._format_preferences(dest_detail.hotel_preferences or {})),
                ("Meal Preferences:", ", ".join(dest_detail.meal_preferences or []) if dest_detail.meal_preferences else "Standard"),
                ("Activities:", ", ".join(dest_detail.activities or []) if dest_detail.activities else "As per itinerary"),
                ("Transportation:", dest_detail.transportation or "As per package"),
                ("Guide Requirements:", dest_detail.guide_requirements or "Not specified"),
                ("Special Notes:", ", ".join(dest_detail.special_notes or []) if dest_detail.special_notes else "None"),
            ]
            
            for label, value in dest_info:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 2
        
        self._apply_sheet_styling(ws)
    
    def _create_complex_pricing_sheet(self, wb: Workbook, quote_data: TravelQuoteData):
        """Create complex pricing sheet with destination-wise breakdown"""
        ws = wb.create_sheet("Pricing Options")
        
        # Header
        ws.merge_cells('A1:H1')
        ws['A1'] = "COMPREHENSIVE PRICING OPTIONS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        
        for option in quote_data.pricing_options:
            # Package header
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = option.get('package_name', 'Package Option')
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            row += 2
            
            # Destination-wise pricing
            for dest in option.get('destinations', []):
                ws[f'A{row}'] = f"📍 {dest['destination']}"
                ws[f'A{row}'].font = Font(bold=True, color="4472C4")
                row += 1
                
                ws[f'A{row}'] = f"  Duration: {dest['nights']} nights"
                row += 1
                
                pricing_items = ['accommodation', 'meals', 'activities', 'transportation', 'guide_services']
                for item in pricing_items:
                    if item in dest and dest[item]:
                        ws[f'B{row}'] = item.replace('_', ' ').title()
                        ws[f'F{row}'] = f"₹ {dest[item]:,.2f}"
                        row += 1
                
                ws[f'B{row}'] = "Subtotal"
                ws[f'F{row}'] = f"₹ {dest['subtotal']:,.2f}"
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'F{row}'].font = Font(bold=True)
                row += 2
            
            # Global costs
            if 'global_costs' in option:
                ws[f'A{row}'] = "🌐 Global Services"
                ws[f'A{row}'].font = Font(bold=True, color="4472C4")
                row += 1
                
                for service, cost in option['global_costs'].items():
                    if cost > 0:
                        ws[f'B{row}'] = service.replace('_', ' ').title()
                        ws[f'F{row}'] = f"₹ {cost:,.2f}"
                        row += 1
                
                row += 1
            
            # Total
            ws[f'A{row}'] = "TOTAL PER PERSON"
            ws[f'F{row}'] = f"₹ {option.get('total_per_person', 0):,.2f}"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'F{row}'].font = Font(bold=True, size=12)
            row += 3
        
        self._apply_sheet_styling(ws)
    
    def _create_simple_summary_sheet(self, wb: Workbook, inquiry: TravelInquiryData, quote_data: TravelQuoteData):
        """Create summary sheet for simple inquiries"""
        ws = wb.active
        ws.title = "Travel Summary"
        
        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Main header
        ws.merge_cells('A1:F1')
        ws['A1'] = "TRAVEL QUOTATION"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        
        # Quote information
        quote_info = [
            ("Quote ID:", quote_data.quote_id),
            ("Date:", datetime.now().strftime("%Y-%m-%d")),
            ("Valid Until:", quote_data.valid_until.strftime("%Y-%m-%d") if quote_data.valid_until else "30 days"),
            ("Number of Options:", len(quote_data.pricing_options))
        ]
        
        for label, value in quote_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Travel details section
        self._add_section_header(ws, row, "TRAVEL DETAILS", header_font, header_fill)
        row += 2
        
        details = [
            ("Number of Travelers:", inquiry.traveler_info.total or "Not specified"),
            ("Destinations:", ", ".join(inquiry.destinations or []) if inquiry.destinations else "Not specified"),
            ("Travel Dates:", self._format_travel_dates(inquiry.travel_dates)),
            ("Duration:", f"{(inquiry.duration or {}).get('total_days', 'Not specified')} days, {(inquiry.duration or {}).get('total_nights', 'Not specified')} nights" if inquiry.duration else "Not specified"),
            ("Departure City:", inquiry.departure_city or "Not specified"),
            ("Budget per Person:", f"₹{inquiry.budget_per_person:,.2f}" if inquiry.budget_per_person else "Not specified"),
        ]
        
        for label, value in details:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Preferences section
        self._add_section_header(ws, row, "PREFERENCES & REQUIREMENTS", header_font, header_fill)
        row += 2
        
        preferences = [
            ("Hotel Preferences:", self._format_preferences(inquiry.global_hotel_preferences or {})),
            ("Meal Preferences:", ", ".join(inquiry.global_meal_preferences or []) if inquiry.global_meal_preferences else "Standard"),
            ("Activities:", ", ".join(inquiry.global_activities or []) if inquiry.global_activities else "As per itinerary"),
            ("Guide Language:", ", ".join(inquiry.guide_language_preferences or []) if inquiry.guide_language_preferences else "English"),
            ("Special Requirements:", self._format_special_requirements(inquiry)),
        ]
        
        for label, value in preferences:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        self._apply_sheet_styling(ws)
    
    def _create_simple_pricing_sheet(self, wb: Workbook, quote_data: TravelQuoteData):
        """Create simple pricing sheet"""
        ws = wb.create_sheet("Pricing Options")
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = "PRICING OPTIONS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        
        for option in quote_data.pricing_options:
            # Option header
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = option.get('package_name', 'Package Option')
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            row += 1
            
            # Pricing items
            pricing_items = ['accommodation', 'transportation', 'meals', 'sightseeing', 'guide_services', 'miscellaneous']
            for item in pricing_items:
                if item in option:
                    ws[f'A{row}'] = item.replace('_', ' ').title()
                    ws[f'E{row}'] = f"₹ {option[item]:,.2f}" if isinstance(option[item], (int, float)) else option[item]
                    row += 1
            
            # Total
            ws[f'A{row}'] = "Total per person"
            ws[f'E{row}'] = f"₹ {option.get('total_per_person', 0):,.2f}"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'E{row}'].font = Font(bold=True)
            row += 3
        
        self._apply_sheet_styling(ws)
    
    def _create_simple_itinerary_sheet(self, wb: Workbook, quote_data: TravelQuoteData):
        """Create simple itinerary sheet"""
        ws = wb.create_sheet("Detailed Itinerary")
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = "DETAILED TRAVEL ITINERARY"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Column headers
        headers = ["Day", "Date", "Destination", "Activities", "Meals", "Accommodation"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Itinerary data
        row = 4
        for day_info in quote_data.itinerary:
            ws.cell(row=row, column=1, value=day_info.get('day', ''))
            ws.cell(row=row, column=2, value=day_info.get('date', ''))
            ws.cell(row=row, column=3, value=day_info.get('destination', ''))
            ws.cell(row=row, column=4, value=day_info.get('activities', ''))
            ws.cell(row=row, column=5, value=day_info.get('meals', ''))
            ws.cell(row=row, column=6, value=day_info.get('accommodation', ''))
            row += 1
        
        self._apply_sheet_styling(ws)
    
    def _create_detailed_itinerary_sheet(self, wb: Workbook, quote_data: TravelQuoteData):
        """Create detailed itinerary sheet for complex inquiries"""
        ws = wb.create_sheet("Detailed Itinerary")
        
        # Header
        ws.merge_cells('A1:H1')
        ws['A1'] = "COMPREHENSIVE TRAVEL ITINERARY"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Column headers
        headers = ["Day", "Date", "Destination", "Activities", "Meals", "Accommodation", "Transportation", "Notes"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Itinerary data
        row = 4
        for day_info in quote_data.itinerary:
            ws.cell(row=row, column=1, value=day_info.get('day', ''))
            ws.cell(row=row, column=2, value=day_info.get('date', ''))
            ws.cell(row=row, column=3, value=day_info.get('destination', ''))
            ws.cell(row=row, column=4, value=day_info.get('activities', ''))
            ws.cell(row=row, column=5, value=day_info.get('meals', ''))
            ws.cell(row=row, column=6, value=day_info.get('accommodation', ''))
            ws.cell(row=row, column=7, value=day_info.get('transportation', ''))
            ws.cell(row=row, column=8, value=day_info.get('special_notes', ''))
            row += 1
        
        self._apply_sheet_styling(ws)
    
    def _create_enhanced_terms_sheet(self, wb: Workbook, quote_data: TravelQuoteData):
        """Create enhanced terms and conditions sheet"""
        ws = wb.create_sheet("Terms & Conditions")
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = "TERMS & CONDITIONS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        
        # Inclusions
        ws[f'A{row}'] = "✅ INCLUSIONS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="008000")
        row += 1
        
        for inclusion in quote_data.inclusions:
            ws[f'A{row}'] = f"• {inclusion}"
            row += 1
        
        row += 2
        
        # Exclusions
        ws[f'A{row}'] = "❌ EXCLUSIONS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
        row += 1
        
        for exclusion in quote_data.exclusions:
            ws[f'A{row}'] = f"• {exclusion}"
            row += 1
        
        row += 2
        
        # General Terms
        ws[f'A{row}'] = "📋 GENERAL TERMS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for i, term in enumerate(quote_data.terms_conditions, 1):
            ws[f'A{row}'] = f"{i}. {term}"
            row += 1
        
        row += 2
        
        # Cancellation Policy
        if quote_data.cancellation_policy:
            ws[f'A{row}'] = "🚫 CANCELLATION POLICY"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FF6600")
            row += 1
            ws[f'A{row}'] = quote_data.cancellation_policy
            row += 1
        
        self._apply_sheet_styling(ws)
    
    def _add_section_header(self, ws, row: int, title: str, font: Font, fill: PatternFill):
        """Add a section header to the worksheet"""
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = font
        ws[f'A{row}'].fill = fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    def _format_travel_dates(self, travel_dates: Optional[Dict]) -> str:
        """Format travel dates for display"""
        if not travel_dates:
            return "Not specified"
        
        try:
            if isinstance(travel_dates, dict):
                start = travel_dates.get('start', '')
                end = travel_dates.get('end', '')
                if start and end:
                    return f"{start} to {end}"
            return str(travel_dates)
        except Exception:
            return "Not specified"
    
    def _format_preferences(self, preferences: Dict) -> str:
        """Format preferences dictionary for display"""
        if not preferences:
            return "Standard"
        if isinstance(preferences, dict):
            formatted = []
            for k, v in (preferences or {}).items():
                if v:
                    formatted.append(f"{k.title()}: {v}")
            return "; ".join(formatted) if formatted else "Standard"
        return str(preferences)
    
    def _format_special_requirements(self, inquiry: TravelInquiryData) -> str:
        """Format all special requirements"""
        requirements = []
        if inquiry.accessibility_requirements:
            requirements.extend(inquiry.accessibility_requirements or [])
        if inquiry.dietary_restrictions:
            requirements.extend(inquiry.dietary_restrictions or [])
        return ", ".join(requirements) if requirements else "None"
    
    def _apply_sheet_styling(self, ws):
        """Apply enhanced styling to worksheet"""
        from openpyxl.worksheet.cell_range import CellRange
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in ws[column_letter]:
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply borders and alignment
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and not isinstance(cell, MergedCell):
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    if not cell.font.bold:
                        cell.font = Font(size=10)

# Export the enhanced class with the original name for compatibility
ExcelQuoteGenerator = EnhancedExcelQuoteGenerator
